from typing import List, Dict
from indox.vector_stores.utils import filter_complex_metadata
from indox.core import Document

class Excel:
    """
    Load an Excel file and extract its sheets as Document objects.

    Parameters:
    - excel_path (str): The path to the Excel file to be loaded.

    Methods:
    - load(): Reads the Excel file, extracts data from each sheet, and creates a dictionary of `Document` objects.

    Returns:
    - Dict[str, List[Document]]: A dictionary where keys are sheet names and values are lists containing `Document` objects with sheet data.

    Raises:
    - FileNotFoundError: If the specified file does not exist.
    - RuntimeError: For any other errors encountered during Excel file processing.
    """

    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        self.sheets: Dict[str, List[Document]] = {}
        self.metadata = {}

    def load(self) -> Dict[str, List[Document]]:
        from openpyxl import load_workbook
        from indox.vector_stores.utils import filter_complex_metadata

        try:
            workbook = load_workbook(filename=self.excel_path, data_only=True)

            self.metadata = {prop: getattr(workbook.properties, prop)
                             for prop in dir(workbook.properties)
                             if not prop.startswith('_') and not callable(getattr(workbook.properties, prop))}

            filtered_excel_reader = filter_complex_metadata([self])[0]
            self.metadata = filtered_excel_reader.metadata

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_data = [[cell.value for cell in row] for row in sheet.iter_rows()]
                document = Document(page_content=sheet_data, **self.metadata)
                self.sheets[sheet_name] = [document]

            for sheet_name in self.sheets:
                self.sheets[sheet_name][0].metadata = self.metadata

            return self.sheets

        except FileNotFoundError:
            raise FileNotFoundError(f"The specified file '{self.excel_path}' does not exist.")
        except Exception as e:
            raise RuntimeError(f"An error occurred while processing the Excel file: {e}")

    def load_and_split(self, splitter, remove_stopwords=False):
        from indox.data_loaders.utils import load_and_process_input
        return load_and_process_input(loader=self.load, splitter=splitter, remove_stopwords=remove_stopwords)
